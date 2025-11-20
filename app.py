"""
Configuration-driven Flask application for reconciliation tool.
All processing logic is now driven by configuration files.
"""

import os
import pandas as pd
import tempfile
import shutil
import smtplib
import threading
import uuid
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from flask import Flask, render_template, request, send_file, jsonify, session, redirect, url_for
from werkzeug.utils import secure_filename
from config import ReconciliationConfig
from processors import ReconciliationProcessor
from rate_tool_integration import run_rate_analysis, save_uploaded_file
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import json

app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = "uploads"
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-key-change-in-production-12345")

# Configure OpenAI API Key (required for root cause analysis)
openai_api_key = os.environ.get("OPENAI_API_KEY")
if not openai_api_key:
    print("WARNING: OPENAI_API_KEY not set. Root cause analysis will not work.")
app.config["OPENAI_API_KEY"] = openai_api_key or ""
# Set as environment variable for use in other modules
if openai_api_key:
    os.environ["OPENAI_API_KEY"] = openai_api_key

os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

# Global instances
config = ReconciliationConfig()
processor = ReconciliationProcessor()
last_output = None
last_rate_report = None

# Batch job tracking
batch_jobs = {}  # {job_id: {status, progress, results, error}}

# Processing state for async background jobs
processing_state = {}  # {session_id: {status, current_index, total, transactions}}
processing_lock = threading.Lock()  # Thread-safe access to processing_state

# Email configuration - with environment variable override
SMTP_HOST = os.environ.get("SMTP_HOST", "smtp.mailjet.com")
SMTP_PORT = int(os.environ.get("SMTP_PORT", 587))
SMTP_USER = os.environ.get("SMTP_USER", "770477fa4a7c9c7c8aac64807c3c69ce")
SMTP_PASS = os.environ.get("SMTP_PASS", "81a49ef6dc5e97dbe4cd67fc95a74fa7")
EMAIL_SENDER = os.environ.get("EMAIL_SENDER", "akshit.mahajan713@gmail.com")
EMAIL_RECIPIENT = os.environ.get("EMAIL_RECIPIENT", "akshit.mahajan0703@gmail.com")
EMAIL_ENABLED = os.environ.get("EMAIL_ENABLED", "true").lower() == "true"

def send_reconciliation_alert(report, transaction_name=""):
    """
    Send email alert when Amount Reconciled falls below 95%

    Args:
        report: Report context with reconciliation metrics
        transaction_name: Name of the transaction (optional)

    Returns:
        bool: True if email sent successfully, False otherwise
    """
    try:
        # Check if email is enabled
        if not EMAIL_ENABLED:
            print("📧 Email alerts disabled - skipping")
            return False

        amount_reconciled = report["summary"]["amount_reconciled_percentage"]

        # Only send email if amount reconciled is below 95%
        if amount_reconciled >= 95:
            print(f"✅ Amount reconciled {amount_reconciled:.2f}% >= 95% - no alert needed")
            return False

        # Prepare email content
        transaction_info = f" - {transaction_name}" if transaction_name else ""
        subject = f"⚠️ Reconciliation Alert{transaction_info}: {amount_reconciled:.2f}%"

        # Create detailed email body
        body = f"""
Reconciliation Alert - Low Match Percentage Detected{transaction_info}

CRITICAL METRICS:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Amount Reconciled: {amount_reconciled:.4f}% ⚠️ (Below 95% threshold)

RECONCILIATION SUMMARY:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Calculated Total (INR): {report["summary"]["total_final_amount_display"]}
• VISA Invoice Total (INR): {report["summary"]["total_visa_amount_display"]}
• Fee Reconciled: {report["summary"]["fee_reconciled_display"]}
• Items Reconciled: {report["summary"]["matched_items"]}/{report["summary"]["total_visa_items"]}
• Amount Match Percentage: {report["summary"]["amount_match_display"]}

ADDITIONAL DETAILS:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Total Fee Mappings: {report["summary"]["total_mappings"]}
• Sheets Analyzed: {report["summary"]["sheet_count"]}

ACTION REQUIRED:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
The reconciliation percentage is below the acceptable threshold of 95%.
Please review the detailed reconciliation report and investigate discrepancies.

This is an automated alert from the Card Reconciliation Tool.
"""

        # Create message
        msg = MIMEMultipart()
        msg["Subject"] = subject
        msg["From"] = EMAIL_SENDER
        msg["To"] = EMAIL_RECIPIENT
        msg.attach(MIMEText(body, "plain"))

        # Send email with timeout
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASS)
            server.send_message(msg)

        print(f"✅ Alert email sent successfully! Amount Reconciled: {amount_reconciled:.2f}%")
        return True

    except smtplib.SMTPAuthenticationError as e:
        print(f"❌ Email authentication failed: {str(e)}")
        return False
    except smtplib.SMTPException as e:
        print(f"❌ SMTP error sending email: {str(e)}")
        return False
    except Exception as e:
        print(f"❌ Error sending email alert: {str(e)}")
        return False

def scan_transaction_folders(base_path):
    """
    Scan base path for transaction subfolders

    Args:
        base_path: Path to transactions folder

    Returns:
        list: List of transaction folder paths
    """
    transaction_folders = []

    if not os.path.exists(base_path):
        return transaction_folders

    for item in os.listdir(base_path):
        item_path = os.path.join(base_path, item)
        if os.path.isdir(item_path):
            transaction_folders.append({
                'name': item,
                'path': item_path
            })

    return sorted(transaction_folders, key=lambda x: x['name'])

def map_files_in_folder(folder_path):
    """
    Automatically map files in a folder to their types

    Args:
        folder_path: Path to transaction folder

    Returns:
        dict: Mapped file paths
    """
    file_mapping = {
        'summary': None,
        'card': None,
        'international': None,
        'domestic': None,
        'dispute': None,
        'invoice': None
    }

    if not os.path.exists(folder_path):
        return file_mapping

    for filename in os.listdir(folder_path):
        filepath = os.path.join(folder_path, filename)

        if not os.path.isfile(filepath):
            continue

        if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
            continue

        filename_lower = filename.lower()

        # Map files based on filename patterns
        if 'summary' in filename_lower:
            file_mapping['summary'] = filepath
        elif 'invoice' in filename_lower:
            file_mapping['invoice'] = filepath
        elif 'card' in filename_lower and 'issuance' in filename_lower:
            file_mapping['card'] = filepath
        elif 'international' in filename_lower:
            file_mapping['international'] = filepath
        elif 'domestic' in filename_lower:
            file_mapping['domestic'] = filepath
        elif 'vrol' in filename_lower or 'dispute' in filename_lower:
            file_mapping['dispute'] = filepath

    return file_mapping

def process_transaction_batch(base_path, job_id=None):
    """
    Process all transaction folders in batch

    Args:
        base_path: Path to transactions folder
        job_id: Optional job ID for progress tracking

    Returns:
        list: List of processing results for each transaction
    """
    results = []
    transaction_folders = scan_transaction_folders(base_path)

    # Update progress: Fetching transactions
    if job_id:
        batch_jobs[job_id]['progress'] = 'fetching'
        batch_jobs[job_id]['total_transactions'] = len(transaction_folders)
        batch_jobs[job_id]['processed'] = 0

    for idx, folder_info in enumerate(transaction_folders):
        folder_name = folder_info['name']
        folder_path = folder_info['path']

        print(f"Processing {folder_name}...")

        # Update progress: Running reconciliation
        if job_id:
            batch_jobs[job_id]['progress'] = 'reconciling'
            batch_jobs[job_id]['current_transaction'] = folder_name
            batch_jobs[job_id]['processed'] = idx

        # Map files
        file_paths = map_files_in_folder(folder_path)

        # Check if we have at least the summary file
        if not file_paths['summary']:
            results.append({
                'transaction_name': folder_name,
                'status': 'failed',
                'error': 'Summary file not found',
                'report': None,
                'email_sent': False
            })
            continue

        try:
            # Run rate analysis
            report = run_rate_analysis(file_paths)

            # Check if email should be sent
            email_sent = False
            if report and report.get("summary"):
                amount_reconciled = report["summary"]["amount_reconciled_percentage"]
                if amount_reconciled < 95:
                    send_reconciliation_alert(report)
                    email_sent = True

            results.append({
                'transaction_name': folder_name,
                'status': 'success',
                'error': None,
                'report': report,
                'email_sent': email_sent
            })

        except Exception as e:
            results.append({
                'transaction_name': folder_name,
                'status': 'failed',
                'error': str(e),
                'report': None,
                'email_sent': False
            })

    # Update progress: Finalizing
    if job_id:
        batch_jobs[job_id]['progress'] = 'finalizing'

    return results

def run_batch_processing_thread(job_id, folder_path):
    """
    Run batch processing in a background thread and update job status

    Args:
        job_id: Unique job identifier
        folder_path: Path to transactions folder
    """
    try:
        # Process the batch
        results = process_transaction_batch(folder_path, job_id)

        # Generate PDF report
        pdf_path = generate_batch_pdf_report(results)

        # Update job status to completed
        batch_jobs[job_id]['status'] = 'completed'
        batch_jobs[job_id]['results'] = results
        batch_jobs[job_id]['pdf_path'] = pdf_path

    except Exception as e:
        # Update job status to failed
        batch_jobs[job_id]['status'] = 'failed'
        batch_jobs[job_id]['error'] = str(e)

def generate_batch_pdf_report(results):
    """
    Generate comprehensive PDF report for batch processing

    Args:
        results: List of processing results

    Returns:
        str: Path to generated PDF file
    """
    output_path = "batch_reconciliation_report.pdf"
    doc = SimpleDocTemplate(output_path, pagesize=A4)

    # Container for PDF elements
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=12,
        spaceBefore=12
    )

    # Title
    title = Paragraph("Batch Reconciliation Report", title_style)
    elements.append(title)

    # Generation date
    date_text = Paragraph(
        f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        styles['Normal']
    )
    elements.append(date_text)
    elements.append(Spacer(1, 20))

    # Summary section
    summary_heading = Paragraph("Executive Summary", heading_style)
    elements.append(summary_heading)

    total_transactions = len(results)
    successful = sum(1 for r in results if r['status'] == 'success')
    failed = sum(1 for r in results if r['status'] == 'failed')
    emails_sent = sum(1 for r in results if r.get('email_sent', False))

    summary_data = [
        ['Metric', 'Value'],
        ['Total Transactions', str(total_transactions)],
        ['Successful', str(successful)],
        ['Failed', str(failed)],
        ['Email Alerts Sent', str(emails_sent)]
    ]

    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 30))

    # Detail for each transaction
    for result in results:
        transaction_heading = Paragraph(
            f"Transaction: {result['transaction_name']}",
            heading_style
        )
        elements.append(transaction_heading)

        if result['status'] == 'failed':
            error_text = Paragraph(
                f"<b>Status:</b> <font color='red'>FAILED</font><br/>"
                f"<b>Error:</b> {result['error']}",
                styles['Normal']
            )
            elements.append(error_text)
            elements.append(Spacer(1, 20))
            continue

        report = result['report']

        # Status with email indicator
        status_color = 'red' if result['email_sent'] else 'green'
        status_text = f"<b>Status:</b> <font color='{status_color}'>SUCCESS"
        if result['email_sent']:
            status_text += " (Email Alert Sent)"
        status_text += "</font>"

        status_para = Paragraph(status_text, styles['Normal'])
        elements.append(status_para)
        elements.append(Spacer(1, 10))

        # Key metrics
        if report and report.get('summary'):
            metrics_data = [
                ['Metric', 'Value'],
                ['Amount Reconciled', report['summary']['amount_reconciled_display']],
                ['Fee Reconciled', report['summary']['fee_reconciled_display']],
                ['Items Reconciled', f"{report['summary']['matched_items']}/{report['summary']['total_visa_items']}"],
                ['Amount Match %', report['summary']['amount_match_display']],
                ['Calculated Total', report['summary']['total_final_amount_display']],
                ['VISA Total', report['summary']['total_visa_amount_display']],
                ['Fee Mappings', str(report['summary']['total_mappings'])]
            ]

            metrics_table = Table(metrics_data, colWidths=[2.5*inch, 2.5*inch])

            # Highlight row if amount reconciled < 95%
            bg_color = colors.HexColor('#FFE6E6') if result['email_sent'] else colors.lightgrey

            metrics_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007BFF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), bg_color),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))

            elements.append(metrics_table)

        elements.append(Spacer(1, 20))

        # Add page break between transactions (except last one)
        if result != results[-1]:
            elements.append(PageBreak())

    # Build PDF
    doc.build(elements)

    return output_path

def format_worksheet(worksheet, header_color='1F4E78', header_font_color='FFFFFF'):
    """
    Apply professional formatting to an Excel worksheet

    Args:
        worksheet: openpyxl worksheet object
        header_color: Hex color for header background (default: dark blue)
        header_font_color: Hex color for header font (default: white)
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Define styles
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
    header_font = Font(bold=True, color=header_font_color, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    # Format header row
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Format data rows
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.alignment = cell_alignment
            cell.border = border

    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        # Set width with some padding
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 for very long text
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Set row height for header
    worksheet.row_dimensions[1].height = 25

def generate_rate_report_excel(report):
    """
    Generate comprehensive Excel report with all reconciliation data

    Args:
        report: Report context with all reconciliation metrics

    Returns:
        str: Path to generated Excel file
    """
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        output_path = "rate_reconciliation_report.xlsx"

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

            # Sheet 1: Summary Metrics
            summary_data = {
                'Metric': [
                    'Amount Reconciled',
                    'Fee Reconciled',
                    'Items Reconciled',
                    'Amount Match Percentage',
                    'Calculated Total (INR)',
                    'VISA Invoice Total (INR)',
                    'Total Fee Mappings',
                    'Sheets Analyzed',
                    'Total VISA Items',
                    'Total Calculated Items',
                    'Matched Items',
                    'Exact Match Items'
                ],
                'Value': [
                    report['summary']['amount_reconciled_display'],
                    report['summary']['fee_reconciled_display'],
                    f"{report['summary']['matched_items']}/{report['summary']['total_visa_items']}",
                    report['summary']['amount_match_display'],
                    report['summary']['total_final_amount_display'],
                    report['summary']['total_visa_amount_display'],
                    report['summary']['total_mappings'],
                    report['summary']['sheet_count'],
                    report['summary']['total_visa_items'],
                    report['summary']['total_calculated_items'],
                    report['summary']['matched_items'],
                    report['summary']['exact_match_items']
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            format_worksheet(writer.sheets['Summary'], header_color='1F4E78')

            # Sheet 2: Card Issuance Data
            if report.get('card'):
                card_data = {
                    'Total Cards Issued': [report['card']['total_cards']]
                }
                card_df = pd.DataFrame(card_data)
                card_df.to_excel(writer, sheet_name='Card Issuance Summary', index=False)
                format_worksheet(writer.sheets['Card Issuance Summary'], header_color='28A745')

                if report['card'].get('monthly_data'):
                    monthly_data = []
                    for entry in report['card']['monthly_data']:
                        monthly_data.append({
                            'Period': entry['period'],
                            'Cards Issued': entry['cards']
                        })
                    monthly_df = pd.DataFrame(monthly_data)
                    monthly_df.to_excel(writer, sheet_name='Card Issuance Detail', index=False)
                    format_worksheet(writer.sheets['Card Issuance Detail'], header_color='28A745')

            # Sheet 3: Transaction Overview
            if report.get('transactions'):
                trans_data = []
                for entry in report['transactions']['entries']:
                    trans_data.append({
                        'Type': entry['label'],
                        'Amount (USD)': entry['amount'],
                        'Volume': entry['volume']
                    })
                trans_df = pd.DataFrame(trans_data)
                trans_df.to_excel(writer, sheet_name='Transaction Overview', index=False)
                format_worksheet(writer.sheets['Transaction Overview'], header_color='007BFF')

            # Sheet 4-N: Detailed Fee Analysis by Sheet
            if report.get('sheets'):
                for idx, sheet in enumerate(report['sheets']):
                    fee_data = []
                    for row in sheet['rows']:
                        fee_data.append({
                            'Fee Type': row['fee_type'],
                            'Rate Chart': row['rate_chart'],
                            'Calculation Method': row['calculation_method'],
                            'Calculated Amount': row['calculated_amount_display'],
                            'Exchange Rate': row['exchange_rate'] if row['exchange_rate'] else 'N/A',
                            'Final Amount (INR)': row['final_amount_display'],
                            'VISA Amount (INR)': row['visa_amount_display'],
                            'Percentage Difference': row['percentage_diff_display'],
                            'Status': row['diff_status']
                        })

                    if fee_data:
                        fee_df = pd.DataFrame(fee_data)
                        sheet_name = sheet['name'][:31]  # Excel sheet name limit is 31 characters
                        fee_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        format_worksheet(writer.sheets[sheet_name], header_color='6F42C1')

            # Sheet: Warnings (if any)
            if report.get('warnings'):
                warnings_data = {
                    'Warnings': report['warnings']
                }
                warnings_df = pd.DataFrame(warnings_data)
                warnings_df.to_excel(writer, sheet_name='Warnings', index=False)
                format_worksheet(writer.sheets['Warnings'], header_color='FFC107', header_font_color='000000')

        return output_path

    except Exception as e:
        print(f"❌ Error generating Excel report: {str(e)}")
        raise

@app.route("/", methods=["GET", "POST"])
def index():
    """Main route - handles both form display and processing"""
    global last_output
    result = None
    error_message = None
    recon_type = request.form.get("recon_type")

    if request.method == "POST":
        try:
            # Validate reconciliation type
            if not config.validate_reconciliation_type(recon_type):
                error_message = "Please select a valid reconciliation type."
            else:
                # Process the reconciliation
                result = processor.process(recon_type, request.files)
                last_output = pd.DataFrame(result) if result else None
                
        except Exception as e:
            error_message = f"Error processing files: {str(e)}"
    
    # Get all available reconciliation types for the form
    reconciliation_types = config.get_all_types()
    
    return render_template("index_dynamic.html", 
                         result=result, 
                         recon_type=recon_type, 
                         error_message=error_message,
                         reconciliation_types=reconciliation_types,
                         result_config=config.RESULT_TABLES.get(recon_type, {}))

@app.route("/api/reconciliation-types")
def get_reconciliation_types():
    """API endpoint to get all reconciliation types configuration"""
    return jsonify(config.get_all_types())

@app.route("/rates-file", methods=["GET", "POST"])
def rates_file():
    """Integrated rates calculator tab"""
    global last_rate_report
    report = None
    error_message = None

    if request.method == "POST":
        temp_dir = tempfile.mkdtemp(prefix="rates_tab_")
        try:
            summary_file = request.files.get("summary_file")
            if not summary_file or summary_file.filename == "":
                raise ValueError("Summary file is required.")

            summary_path = save_uploaded_file(summary_file, temp_dir)
            card_path = save_uploaded_file(request.files.get("card_file"), temp_dir)
            international_path = save_uploaded_file(request.files.get("international_file"), temp_dir)
            domestic_path = save_uploaded_file(request.files.get("domestic_file"), temp_dir)
            dispute_path = save_uploaded_file(request.files.get("dispute_file"), temp_dir)
            invoice_path = save_uploaded_file(request.files.get("invoice_file"), temp_dir)

            file_paths = {
                "summary": summary_path,
                "card": card_path,
                "international": international_path,
                "domestic": domestic_path,
                "dispute": dispute_path,
                "invoice": invoice_path
            }

            report = run_rate_analysis(file_paths)

            # Store report globally for download
            last_rate_report = report

            # Send email alert if Amount Reconciled is below 95%
            if report and report.get("summary"):
                send_reconciliation_alert(report)

        except Exception as exc:
            error_message = str(exc)
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    return render_template("rates_tab.html", report=report, error_message=error_message)

@app.route("/upload-transactions", methods=["POST"])
def upload_transactions():
    """Upload ZIP file and extract transactions to workspace"""
    try:
        # Get ZIP file
        zip_file = request.files.get("transactions_zip")

        if not zip_file or zip_file.filename == "":
            return jsonify({"error": "Please upload a ZIP file"}), 400

        # Create unique session ID
        session_id = str(uuid.uuid4())

        # Extract to temp folder with session ID
        temp_dir = os.path.join(tempfile.gettempdir(), f"batch_{session_id}")
        os.makedirs(temp_dir, exist_ok=True)

        # Save and extract ZIP
        zip_path = os.path.join(temp_dir, "upload.zip")
        zip_file.save(zip_path)

        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)

        # Remove the zip file after extraction
        os.remove(zip_path)

        # Find the extracted folder (might be nested)
        extracted_folders = [f for f in os.listdir(temp_dir) if os.path.isdir(os.path.join(temp_dir, f))]

        if len(extracted_folders) == 1:
            transactions_folder = os.path.join(temp_dir, extracted_folders[0])
        else:
            transactions_folder = temp_dir

        # Scan for transaction folders
        transaction_folders = scan_transaction_folders(transactions_folder)

        if not transaction_folders:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return jsonify({"error": "No transaction folders found in the ZIP file"}), 400

        # Build metadata for each transaction
        transactions_data = []
        for folder_info in transaction_folders:
            folder_path = folder_info['path']
            file_mapping = map_files_in_folder(folder_path)

            transactions_data.append({
                'name': folder_info['name'],
                'path': folder_path,
                'files': file_mapping,
                'status': 'pending',
                'has_summary': file_mapping.get('summary') is not None,
                'has_invoice': file_mapping.get('invoice') is not None,
                'file_count': sum(1 for v in file_mapping.values() if v)
            })

        # Store in session
        session['batch_session_id'] = session_id
        session['temp_dir'] = temp_dir
        session['transactions'] = transactions_data
        session['upload_time'] = datetime.now().isoformat()
        session['auto_process_pending'] = True  # Flag to trigger auto-processing
        session.modified = True

        return jsonify({
            "status": "success",
            "session_id": session_id,
            "transaction_count": len(transactions_data)
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/transaction-browser")
def transaction_browser():
    """Display transaction browser workspace"""
    # Check if user has uploaded files
    if 'batch_session_id' not in session or 'transactions' not in session:
        # No active session - redirect to upload page
        return redirect(url_for('rates_file_automated'))

    transactions = session.get('transactions', [])
    session_id = session.get('batch_session_id', '')

    # If no transactions, redirect to upload page
    if not transactions:
        return redirect(url_for('rates_file_automated'))

    # If auto-process is pending, redirect to processing page
    if session.get('auto_process_pending'):
        return redirect(url_for('processing_page'))

    return render_template('transaction_browser.html',
                         transactions=transactions,
                         session_id=session_id)

@app.route("/processing-page")
def processing_page():
    """Show processing page with progress"""
    transactions = session.get('transactions', [])

    if not transactions:
        return redirect(url_for('rates_file_automated'))

    # Count transactions that need processing
    total_count = sum(1 for t in transactions if t.get('status') == 'pending' and t.get('has_summary'))

    return render_template('processing_page.html', total_count=total_count)

def process_single_transaction(idx, transaction, temp_dir):
    """
    Worker function to process a single transaction

    Args:
        idx: Transaction index
        transaction: Transaction dictionary
        temp_dir: Temporary directory for storing reports

    Returns:
        dict: Processing result with index, status, and metadata
    """
    try:
        # Update processing state
        with processing_lock:
            transaction['status'] = 'processing'

        # Run rate analysis
        file_paths = transaction['files']
        report = run_rate_analysis(file_paths)

        # Check email alert
        email_sent = False
        if report and report.get("summary"):
            amount_reconciled = report["summary"]["amount_reconciled_percentage"]
            if amount_reconciled < 95:
                email_sent = send_reconciliation_alert(report, transaction_name=transaction.get('name', ''))

        # Save report to file
        report_filename = f"report_{idx}.json"
        report_path = os.path.join(temp_dir, report_filename)
        with open(report_path, 'w') as f:
            json.dump(report, f)

        # Update transaction status
        with processing_lock:
            transaction['status'] = 'completed'
            transaction['report_file'] = report_filename
            transaction['email_sent'] = email_sent
            transaction['amount_reconciled'] = report['summary']['amount_reconciled_display']

        return {
            'index': idx,
            'status': 'success',
            'email_sent': email_sent,
            'transaction_name': transaction.get('name', '')
        }

    except Exception as e:
        with processing_lock:
            transaction['status'] = 'failed'
            transaction['error'] = str(e)

        return {
            'index': idx,
            'status': 'failed',
            'error': str(e),
            'transaction_name': transaction.get('name', '')
        }

@app.route("/execute-batch-processing", methods=["POST"])
def execute_batch_processing():
    """Execute batch processing with parallel processing for faster execution"""
    try:
        transactions = session.get('transactions', [])
        temp_dir = session.get('temp_dir')

        # Find all pending transactions
        pending_tasks = []
        for idx, txn in enumerate(transactions):
            if txn.get('status') == 'pending' and txn.get('has_summary'):
                pending_tasks.append((idx, txn))

        if not pending_tasks:
            return jsonify({'status': 'completed', 'message': 'No pending transactions to process'})

        # Initialize processing state
        session_id = session.get('session_id', str(uuid.uuid4()))
        with processing_lock:
            processing_state[session_id] = {
                'status': 'processing',
                'current_index': 0,
                'total': len(pending_tasks),
                'completed': 0
            }

        # Process transactions in parallel (max 2 workers for free tier memory constraints)
        max_workers = min(2, len(pending_tasks))
        results = []

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # Submit all tasks
            future_to_idx = {
                executor.submit(process_single_transaction, idx, txn, temp_dir): idx
                for idx, txn in pending_tasks
            }

            # Process results as they complete
            for future in as_completed(future_to_idx):
                result = future.result()
                results.append(result)

                # Update processing state
                with processing_lock:
                    if session_id in processing_state:
                        processing_state[session_id]['completed'] += 1
                        processing_state[session_id]['current_index'] = result['index']

        # Clear processing state
        with processing_lock:
            if session_id in processing_state:
                processing_state[session_id]['status'] = 'completed'

        # Update session
        session['transactions'] = transactions
        session.pop('auto_process_pending', None)
        session.modified = True

        return jsonify({
            'status': 'completed',
            'message': f'Processed {len(results)} transactions',
            'results': results
        })

    except Exception as e:
        print(f"❌ Batch processing error: {str(e)}")
        return jsonify({'status': 'failed', 'error': str(e)}), 500

@app.route("/batch-processing-status")
def batch_processing_status():
    """Get current processing status"""
    return jsonify({
        'current_index': session.get('processing_index', -1),
        'current_transaction': session.get('processing_transaction', ''),
        'status': 'processing' if 'processing_index' in session else 'completed'
    })

@app.route("/process-workspace-transactions", methods=["POST"])
def process_workspace_transactions():
    """Process selected transactions from workspace with parallel processing"""
    try:
        data = request.get_json()
        indices = data.get('transaction_indices', [])

        transactions = session.get('transactions', [])
        temp_dir = session.get('temp_dir')

        # Prepare tasks for parallel processing
        tasks = []
        for idx in indices:
            if idx < len(transactions):
                tasks.append((idx, transactions[idx]))

        if not tasks:
            return jsonify({'results': []})

        # Process transactions in parallel (max 2 workers for free tier)
        max_workers = min(2, len(tasks))
        results = []

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # Submit all tasks
            future_to_idx = {
                executor.submit(process_single_transaction, idx, txn, temp_dir): idx
                for idx, txn in tasks
            }

            # Collect results as they complete
            for future in as_completed(future_to_idx):
                result = future.result()
                results.append(result)

        # Update session
        session['transactions'] = transactions
        session.pop('auto_process_pending', None)
        session.modified = True

        return jsonify({'results': results})

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/workspace-result/<int:index>")
def workspace_result(index):
    """View detailed results for a specific workspace transaction"""
    transactions = session.get('transactions', [])
    temp_dir = session.get('temp_dir')

    if index >= len(transactions):
        return render_template("error.html", error="Transaction not found"), 404

    transaction = transactions[index]

    if transaction['status'] != 'completed':
        return render_template("error.html",
                             error=f"Transaction not processed yet. Status: {transaction['status']}"), 400

    # Load report from file
    report_filename = transaction.get('report_file')
    if not report_filename:
        return render_template("error.html", error="Report data not available"), 404

    import json
    report_path = os.path.join(temp_dir, report_filename)

    if not os.path.exists(report_path):
        return render_template("error.html", error="Report file not found"), 404

    try:
        with open(report_path, 'r') as f:
            report = json.load(f)
    except Exception as e:
        return render_template("error.html", error=f"Error loading report: {str(e)}"), 500

    return render_template('rates_tab.html',
                         report=report,
                         transaction_name=transaction['name'],
                         email_sent=transaction.get('email_sent', False))

@app.route("/clear-auto-process-flag", methods=["POST"])
def clear_auto_process_flag():
    """Clear the auto-process flag to prevent re-triggering"""
    session.pop('auto_process_pending', None)
    session.modified = True
    return jsonify({"status": "cleared"})

@app.route("/clear-workspace", methods=["POST"])
def clear_workspace():
    """Clear workspace and delete temporary files"""
    try:
        temp_dir = session.get('temp_dir')

        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)

        # Clear session data
        session.pop('batch_session_id', None)
        session.pop('temp_dir', None)
        session.pop('transactions', None)
        session.pop('upload_time', None)
        session.pop('auto_process_pending', None)
        session.modified = True

        return jsonify({"status": "cleared"})

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/rates-file-automated", methods=["GET"])
def rates_file_automated():
    """Automated batch processing with ZIP upload - Always show upload page"""
    # Clear any old session data when visiting the upload page
    session.pop('batch_session_id', None)
    session.pop('temp_dir', None)
    session.pop('transactions', None)
    session.pop('upload_time', None)
    session.pop('auto_process_pending', None)
    session.modified = True

    return render_template("rates_tab_automated.html")

@app.route("/clear-batch-results")
def clear_batch_results():
    """Clear stored batch results and redirect to automated batch page"""
    session.pop('batch_results', None)
    session.pop('batch_folder_path', None)
    session.modified = True
    return jsonify({"status": "cleared"})

@app.route("/batch-transaction-detail/<transaction_name>")
def batch_transaction_detail(transaction_name):
    """Display detailed view for a specific transaction from batch processing"""
    # Retrieve batch results from session
    batch_results = session.get('batch_results', [])

    # Find the transaction in batch results
    transaction_result = None
    for result in batch_results:
        if result['transaction_name'] == transaction_name:
            transaction_result = result
            break

    if not transaction_result:
        return render_template("error.html", error="Transaction not found. Please run batch analysis again."), 404

    if transaction_result['status'] == 'failed':
        return render_template("error.html",
                             error=f"Transaction '{transaction_name}' failed: {transaction_result.get('error', 'Unknown error')}"), 500

    # Get the full report for this transaction
    report = transaction_result.get('report')

    if not report:
        return render_template("error.html", error="Report data not available"), 404

    # Render the same detailed view as manual upload
    return render_template("batch_transaction_detail.html",
                         report=report,
                         transaction_name=transaction_name,
                         email_sent=transaction_result.get('email_sent', False))

@app.route("/download-batch-pdf")
def download_batch_pdf():
    """Download batch processing PDF report"""
    try:
        pdf_path = "batch_reconciliation_report.pdf"
        if os.path.exists(pdf_path):
            return send_file(pdf_path, as_attachment=True, download_name="batch_reconciliation_report.pdf")
        else:
            return "No batch report available to download.", 404
    except Exception as e:
        return f"Error downloading report: {str(e)}", 500

@app.route("/download")
def download():
    """Download reconciliation results as Excel file"""
    global last_output
    if last_output is not None:
        path = "reconciliation_output.xlsx"
        last_output.to_excel(path, index=False)
        return send_file(path, as_attachment=True)
    return "No reconciliation results available to download.", 404

@app.route("/download-rate-report")
def download_rate_report():
    """Download comprehensive rate reconciliation report as Excel file"""
    global last_rate_report
    if last_rate_report is not None:
        try:
            path = generate_rate_report_excel(last_rate_report)
            return send_file(path, as_attachment=True, download_name="rate_reconciliation_report.xlsx")
        except Exception as e:
            return f"Error generating report: {str(e)}", 500
    return "No rate analysis results available to download.", 404

@app.route("/start-batch-analysis", methods=["POST"])
def start_batch_analysis():
    """Start batch analysis in background and return job ID"""
    try:
        data = request.get_json()
        folder_path = data.get("folder_path", "").strip()

        if not folder_path:
            return jsonify({"error": "Please enter a folder path."}), 400

        if not os.path.exists(folder_path):
            return jsonify({"error": f"Folder path does not exist: {folder_path}"}), 400

        if not os.path.isdir(folder_path):
            return jsonify({"error": "Path must be a directory."}), 400

        # Generate unique job ID
        job_id = str(uuid.uuid4())

        # Initialize job tracking
        batch_jobs[job_id] = {
            'status': 'processing',
            'progress': 'initializing',
            'folder_path': folder_path,
            'total_transactions': 0,
            'processed': 0,
            'current_transaction': '',
            'results': None,
            'error': None
        }

        # Start processing in background thread
        thread = threading.Thread(target=run_batch_processing_thread, args=(job_id, folder_path))
        thread.daemon = True
        thread.start()

        return jsonify({
            "job_id": job_id,
            "status": "started"
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/batch-progress/<job_id>")
def batch_progress(job_id):
    """Get progress status for a batch job"""
    job = batch_jobs.get(job_id)

    if not job:
        return jsonify({"error": "Job not found"}), 404

    response = {
        "status": job['status'],
        "progress": job['progress']
    }

    if job['status'] == 'processing':
        response['total_transactions'] = job.get('total_transactions', 0)
        response['processed'] = job.get('processed', 0)
        response['current_transaction'] = job.get('current_transaction', '')

    elif job['status'] == 'completed':
        response['results'] = job['results']

    elif job['status'] == 'failed':
        response['error'] = job.get('error', 'Unknown error')

    return jsonify(response)

@app.route("/save-batch-results/<job_id>", methods=["POST"])
def save_batch_results(job_id):
    """Save batch results to session for viewing"""
    job = batch_jobs.get(job_id)

    if not job:
        return jsonify({"error": "Job not found"}), 404

    if job['status'] != 'completed':
        return jsonify({"error": "Job not completed"}), 400

    # Store results in session
    session['batch_results'] = job['results']
    session['batch_folder_path'] = job['folder_path']
    session.modified = True

    return jsonify({"status": "saved"})

@app.route("/health")
def health_check():
    """Health check endpoint"""
    return jsonify({
        "status": "healthy",
        "available_types": list(config.get_all_types().keys()),
        "version": "2.0-config-driven"
    })

@app.route("/test-config")
def test_config():
    """Test configuration endpoint - check if environment variables are set"""
    openai_key_set = bool(os.environ.get("OPENAI_API_KEY"))
    secret_key_set = bool(app.config.get("SECRET_KEY"))

    return jsonify({
        "status": "ok",
        "openai_api_key_configured": openai_key_set,
        "secret_key_configured": secret_key_set,
        "python_version": f"Python {'.'.join(map(str, __import__('sys').version_info[:3]))}",
        "openai_key_length": len(os.environ.get("OPENAI_API_KEY", "")) if openai_key_set else 0
    })

@app.errorhandler(404)
def not_found_error(error):
    return render_template("error.html", error="Page not found"), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template("error.html", error="Internal server error"), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)

